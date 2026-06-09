# -*- coding: utf-8 -*-
"""
Rapor endpoint'leri — PDF ve Excel export
"""
import io, json
from datetime import datetime
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional

router = APIRouter()

CATEGORY_TR = {
    'ALISVERIS': 'Alisveris', 'BANKA_KESINTISI': 'Banka Kes.', 'DIGER': 'Diger',
    'EGITIM': 'Egitim', 'EGLENCE': 'Eglence', 'FATURA': 'Fatura',
    'MARKET': 'Market', 'NAKIT_ISLEMLERI': 'Nakit', 'SAGLIK': 'Saglik',
    'SEYAHAT': 'Seyahat', 'ULASIM': 'Ulasim', 'VIRMAN': 'Virman', 'YEME_ICME': 'Yeme-Icme',
}


def _tr(s) -> str:
    """Turkce karakterleri ASCII'ye donustur (reportlab uyumlulugu)."""
    if not isinstance(s, str):
        s = str(s) if s is not None else ''
    for t, e in [('ç','c'),('ş','s'),('ğ','g'),('ü','u'),('ö','o'),('ı','i'),
                 ('İ','I'),('Ç','C'),('Ş','S'),('Ğ','G'),('Ü','U'),('Ö','O')]:
        s = s.replace(t, e)
    return s


def _make_category_chart(cats: dict) -> io.BytesIO:
    """Kategori bazli harcama yatay bar grafigi."""
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import numpy as np
    except ImportError:
        return None

    data = [(CATEGORY_TR.get(k, k), v) for k, v in cats.items() if v and v > 0]
    if not data:
        return None
    data.sort(key=lambda x: x[1])
    data = data[-8:]  # En yüksek 8

    labels = [d[0] for d in data]
    values = [d[1] for d in data]

    fig, ax = plt.subplots(figsize=(6.5, max(2.5, len(labels) * 0.4)))
    colors = ['#166534', '#16a34a', '#22c55e', '#4ade80', '#86efac',
              '#f59e0b', '#ef4444', '#84cc16']
    bars = ax.barh(labels, values, color=[colors[i % len(colors)] for i in range(len(labels))],
                   height=0.6, edgecolor='white', linewidth=0.5)

    ax.set_xlabel('Tutar (TL)', fontsize=9)
    ax.set_title('Kategori Bazli Harcama Dagilimi', fontsize=10, fontweight='bold', pad=8)
    ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x/1000:.0f}K' if x >= 1000 else f'{x:.0f}'))
    ax.tick_params(axis='both', labelsize=8)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    for bar, val in zip(bars, values):
        ax.text(bar.get_width() + max(values) * 0.01, bar.get_y() + bar.get_height() / 2,
                f'{val:,.0f}', va='center', ha='left', fontsize=7.5, color='#374151')

    fig.patch.set_facecolor('#f9fafb')
    ax.set_facecolor('#f9fafb')
    plt.tight_layout(pad=1.0)

    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight', facecolor='#f9fafb')
    plt.close(fig)
    buf.seek(0)
    return buf


def _make_trend_chart(trend: list, forecast: float) -> io.BytesIO:
    """Aylik gelir/gider trend + tahmin grafigi."""
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import numpy as np
    except ImportError:
        return None

    if not trend:
        return None

    months = [t.get('month', '')[-5:] for t in trend]
    income  = [t.get('income', 0) for t in trend]
    expense = [t.get('expense', 0) for t in trend]

    # Tahmin ayini ekle
    if forecast and forecast > 0:
        months.append('Tahmin')
        income.append(0)
        expense.append(forecast)

    x = range(len(months))
    w = 0.35

    fig, ax = plt.subplots(figsize=(6.5, 3.0))
    bars1 = ax.bar([i - w/2 for i in x], income,  w, label='Gelir', color='#22c55e', edgecolor='white')
    bars2 = ax.bar([i + w/2 for i in x], expense, w, label='Gider', color='#166534', edgecolor='white')
    # Tahmin çubuğuna ayrı alpha uygula
    if forecast and forecast > 0 and len(bars2) > 0:
        bars2[-1].set_alpha(0.55)
        bars2[-1].set_facecolor('#f59e0b')  # tahmin farklı renk

    ax.set_xticks(list(x))
    ax.set_xticklabels(months, fontsize=8, rotation=15)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f'{y/1000:.0f}K' if y >= 1000 else f'{y:.0f}'))
    ax.tick_params(axis='y', labelsize=8)
    ax.set_title('Aylik Gelir / Gider Trendi', fontsize=10, fontweight='bold', pad=8)
    ax.legend(fontsize=8, loc='upper left')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    if forecast and forecast > 0:
        ax.axvline(x=len(months) - 1.5, color='#f59e0b', linestyle='--', linewidth=1, alpha=0.7)
        ax.text(len(months) - 1.4, ax.get_ylim()[1] * 0.95,
                'Tahmin', fontsize=7.5, color='#d97706', va='top')

    fig.patch.set_facecolor('#f9fafb')
    ax.set_facecolor('#f9fafb')
    plt.tight_layout(pad=1.0)

    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight', facecolor='#f9fafb')
    plt.close(fig)
    buf.seek(0)
    return buf


def _make_probs_chart(probs: dict) -> io.BytesIO:
    """AI karar olasilik dagilimi grafigi."""
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
    except ImportError:
        return None

    if not probs:
        return None

    LABEL_TR = {
        'HIGH_DEBT_RISK':  'Riskli',
        'NEEDS_BUDGETING': 'Butceleme Yapmali',
        'GREAT_SAVER':     'Tasarrufcu',
    }
    COLOR_MAP = {
        'HIGH_DEBT_RISK':  '#ef4444',
        'NEEDS_BUDGETING': '#f59e0b',
        'GREAT_SAVER':     '#22c55e',
    }

    labels = [LABEL_TR.get(k, k) for k in probs]
    values = [v * 100 for v in probs.values()]
    colors = [COLOR_MAP.get(k, '#6b7280') for k in probs]

    fig, ax = plt.subplots(figsize=(4.0, 2.2))
    bars = ax.barh(labels, values, color=colors, height=0.5, edgecolor='white')
    ax.set_xlim(0, 105)
    ax.set_xlabel('Olasilik (%)', fontsize=8)
    ax.set_title('AI Karar Dagilimi', fontsize=9, fontweight='bold', pad=6)
    ax.tick_params(axis='both', labelsize=8)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    for bar, val in zip(bars, values):
        ax.text(bar.get_width() + 1, bar.get_y() + bar.get_height() / 2,
                f'%{val:.0f}', va='center', ha='left', fontsize=8, fontweight='bold')

    fig.patch.set_facecolor('#f9fafb')
    ax.set_facecolor('#f9fafb')
    plt.tight_layout(pad=0.8)

    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight', facecolor='#f9fafb')
    plt.close(fig)
    buf.seek(0)
    return buf


class ReportRequest(BaseModel):
    result: dict
    mode: str = "real"
    persona_name: Optional[str] = None


def _generate_pdf(result: dict, mode: str, persona_name: Optional[str]) -> bytes:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.colors import HexColor, black, white
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Image as RLImage
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        raise RuntimeError("reportlab yuklu degil: pip install reportlab")

    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.8*cm, bottomMargin=1.8*cm,
                             leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()

    NAVY   = HexColor('#1e3a5f')
    GREEN  = HexColor('#16a34a')
    RED    = HexColor('#dc2626')
    YELLOW = HexColor('#d97706')
    LGRAY  = HexColor('#f8fafc')
    MGRAY  = HexColor('#e5e7eb')

    title_style = ParagraphStyle('Title',  parent=styles['Title'],   fontSize=18, textColor=NAVY, spaceAfter=3, leading=22)
    h1_style    = ParagraphStyle('H1',     parent=styles['Heading1'],fontSize=13, textColor=NAVY, spaceAfter=4, spaceBefore=6)
    body_style  = ParagraphStyle('Body',   parent=styles['Normal'],  fontSize=9.5, leading=14, spaceAfter=5)
    small_style = ParagraphStyle('Small',  parent=styles['Normal'],  fontSize=7.5, textColor=HexColor('#6b7280'), spaceAfter=4)
    warn_style  = ParagraphStyle('Warn',   parent=styles['Normal'],  fontSize=8.5, textColor=RED, spaceAfter=4)

    story = []
    now   = datetime.now().strftime("%d.%m.%Y %H:%M")
    s     = result.get('summary', {})
    prof  = result.get('user_profile', {})
    coach = result.get('ai_coach', {})
    fore  = result.get('ai_forecast', {})
    shap  = result.get('shap_features', [])
    cats  = result.get('category_breakdown', {})
    trend = result.get('monthly_trend', [])
    probs = coach.get('class_probabilities', {})

    # ── Başlık ───────────────────────────────────────────────────────
    story.append(Paragraph("FinWise — Finansal Analiz Raporu", title_style))
    story.append(Paragraph(
        f"Olusturulma: {now}  |  Mod: {'Demo' if mode == 'demo' else 'Kisisel Analiz'}",
        small_style))
    if mode == 'demo' and persona_name:
        story.append(Paragraph(f"Demo Profil: {_tr(persona_name)} — Sentetik ve anonim ornek veri.", small_style))
    story.append(HRFlowable(width="100%", thickness=1.5, color=NAVY))
    story.append(Spacer(1, 0.3*cm))

    # ── Mali Özet ────────────────────────────────────────────────────
    story.append(Paragraph("Mali Ozet", h1_style))
    summary_data = [
        ["Metrik", "Deger"],
        ["Toplam Gelir",    f"{s.get('total_income',0):,.0f} TL"],
        ["Toplam Gider",    f"{s.get('total_expense',0):,.0f} TL"],
        ["Net Nakit Akisi", f"{s.get('net_cashflow',0):+,.0f} TL"],
        ["Tasarruf Orani",  f"%{s.get('savings_rate_pct',0):.1f}"],
        ["Toplam Islem",    f"{s.get('txn_count',0)}"],
        ["Kategorize Edilen", f"{s.get('categorized_count',0)}"],
    ]
    t = Table(summary_data, colWidths=[8*cm, 7*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), NAVY),
        ('TEXTCOLOR',  (0,0), (-1,0), white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME',   (0,1), (0,-1), 'Helvetica-Bold'),
        ('TEXTCOLOR',  (0,1), (0,-1), NAVY),
        ('FONTSIZE',   (0,0), (-1,-1), 9.5),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [white, LGRAY]),
        ('GRID',       (0,0), (-1,-1), 0.3, MGRAY),
        ('PADDING',    (0,0), (-1,-1), 6),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [white, LGRAY]),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.4*cm))

    # ── Kategori Grafiği ─────────────────────────────────────────────
    cat_buf = _make_category_chart(cats)
    if cat_buf:
        story.append(Paragraph("Kategori Bazli Harcama Dagilimi", h1_style))
        img = RLImage(cat_buf, width=15*cm, height=max(4*cm, min(8*cm, len([v for v in cats.values() if v > 0]) * 0.55*cm)))
        story.append(img)
        story.append(Spacer(1, 0.3*cm))
    else:
        # Grafik yoksa tablo goster
        story.append(Paragraph("Kategori Bazli Harcama", h1_style))
        cat_rows = [["Kategori", "Tutar (TL)"]] + [
            [CATEGORY_TR.get(k, k), f"{v:,.0f}"]
            for k, v in sorted(cats.items(), key=lambda x: -x[1]) if v > 0
        ]
        if len(cat_rows) > 1:
            ct = Table(cat_rows, colWidths=[9*cm, 6*cm])
            ct.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), NAVY), ('TEXTCOLOR', (0,0), (-1,0), white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,-1), 9),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [white, LGRAY]),
                ('GRID', (0,0), (-1,-1), 0.3, MGRAY), ('PADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(ct)
        story.append(Spacer(1, 0.3*cm))

    # ── Aylık Trend Grafiği ──────────────────────────────────────────
    trend_buf = _make_trend_chart(trend, fore.get('next_month_expected_expense', 0))
    if trend_buf:
        story.append(Paragraph("Aylik Gelir / Gider Trendi + Tahmin", h1_style))
        story.append(RLImage(trend_buf, width=15*cm, height=5.5*cm))
        story.append(Spacer(1, 0.3*cm))

    # ── Finansal Profil ──────────────────────────────────────────────
    story.append(Paragraph("Finansal Profil", h1_style))
    profile_data = [
        ["Ozellik", "Deger"],
        ["Finansal Segment",   _tr(prof.get('cluster_name', '-'))],
        ["Nakit Akisi",        _tr(prof.get('cash_flow', '-'))],
        ["Tasarruf Egilimi",   _tr(prof.get('savings_trend', '-'))],
        ["Harcama Ritmi",      _tr(prof.get('spending_rhythm', '-'))],
        ["Baskin Kategori",    CATEGORY_TR.get(prof.get('dominant_category',''), prof.get('dominant_category','-'))],
    ]
    t2 = Table(profile_data, colWidths=[8*cm, 7*cm])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), NAVY), ('TEXTCOLOR', (0,0), (-1,0), white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME', (0,1), (0,-1), 'Helvetica-Bold'), ('TEXTCOLOR', (0,1), (0,-1), NAVY),
        ('FONTSIZE', (0,0), (-1,-1), 9.5),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [white, LGRAY]),
        ('GRID', (0,0), (-1,-1), 0.3, MGRAY), ('PADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t2)
    story.append(Spacer(1, 0.4*cm))

    # ── AI Teşhis + Olasılık Grafiği ─────────────────────────────────
    story.append(Paragraph("AI Finansal Teshis", h1_style))
    label = _tr(coach.get('label_display', coach.get('action_label', '-')))
    label_color_hex = '#dc2626' if 'Riskli' in label or 'Riski' in label else (
                      '#d97706' if 'Butce' in label or 'Butceleme' in label else '#16a34a')
    story.append(Paragraph(
        f"<b>Teshis:</b> <font color='{label_color_hex}'><b>{label}</b></font>",
        body_style))
    story.append(Paragraph(
        f"Gelecek Ay Tahmini: <b>{fore.get('next_month_expected_expense', 0):,.0f} TL</b>"
        f" ({_tr(fore.get('trend',''))})",
        body_style))
    story.append(Paragraph(_tr(coach.get('personalized_message', '')), body_style))

    probs_buf = _make_probs_chart(probs)
    if probs_buf:
        story.append(RLImage(probs_buf, width=9*cm, height=4*cm))
    elif probs:
        pdata = [["Sinif", "Olasilik"]] + [[_tr(k), f"%{v*100:.0f}"] for k, v in probs.items()]
        pt = Table(pdata, colWidths=[9*cm, 6*cm])
        pt.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), NAVY), ('TEXTCOLOR', (0,0), (-1,0), white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,-1), 9),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [white, LGRAY]),
            ('GRID', (0,0), (-1,-1), 0.3, MGRAY), ('PADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(pt)
    story.append(Spacer(1, 0.4*cm))

    # ── SHAP Faktörleri ──────────────────────────────────────────────
    if shap:
        story.append(Paragraph("Karari Etkileyen Faktorler (SHAP)", h1_style))
        sdata = [["Faktor", "Deger", "Etki"]]
        for f in shap:
            direction = "Riski Artirdi" if f.get('direction') == 'risk' else (
                        "Riski Azaltti" if f.get('direction') == 'safe' else "Notr")
            sdata.append([_tr(f.get('display','')),
                          f"{f.get('value',0):,.1f} {f.get('unit','')}",
                          direction])
        st = Table(sdata, colWidths=[7.5*cm, 4*cm, 3.5*cm])
        st.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), NAVY), ('TEXTCOLOR', (0,0), (-1,0), white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,-1), 9),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [white, LGRAY]),
            ('GRID', (0,0), (-1,-1), 0.3, MGRAY), ('PADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(st)
        story.append(Spacer(1, 0.2*cm))
        story.append(Paragraph(
            "Bu aciklama SHAP tabanli model yorumlamasidir. Nedensellik kaniti degil, karar katki analizidir.",
            small_style))

    # ── Uyarı ────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=MGRAY))
    story.append(Paragraph(
        "Bu rapor bilgilendirme ve karar destek amaciyla uretilmistir. Kesin finansal tavsiye degildir.",
        warn_style))

    doc.build(story)
    return buf.getvalue()


def _generate_excel(result: dict, mode: str) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl yuklu degil: pip install openpyxl")

    wb = Workbook()
    NAVY_HEX = "1e3a5f"
    LGRAY_HEX = "f0f4f8"

    def hdr(ws, row, col, value):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=NAVY_HEX)
        cell.alignment = Alignment(horizontal="center")
        return cell

    ws1 = wb.active
    ws1.title = "Ozet"
    s = result.get('summary', {})
    hdr(ws1, 1, 1, "Metrik"); hdr(ws1, 1, 2, "Deger")
    for i, (label, value) in enumerate([
        ("Toplam Gelir (TL)",    s.get('total_income', 0)),
        ("Toplam Gider (TL)",    s.get('total_expense', 0)),
        ("Net Nakit Akisi (TL)", s.get('net_cashflow', 0)),
        ("Tasarruf Orani (%)",   s.get('savings_rate_pct', 0)),
        ("Toplam Islem",         s.get('txn_count', 0)),
        ("Kategorize Edilen",    s.get('categorized_count', 0)),
        ("Finansal Segment",     result.get('user_profile', {}).get('cluster_name', '-')),
        ("AI Tespiti",           result.get('ai_coach', {}).get('label_display', '-')),
        ("Gelecek Ay Tahmini",   result.get('ai_forecast', {}).get('next_month_expected_expense', 0)),
    ], start=2):
        ws1.cell(row=i, column=1, value=label)
        ws1.cell(row=i, column=2, value=value)
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 20

    ws2 = wb.create_sheet("Kategori Dagilimi")
    hdr(ws2, 1, 1, "Kategori"); hdr(ws2, 1, 2, "Tutar (TL)")
    for i, (k, v) in enumerate(sorted(result.get('category_breakdown', {}).items(), key=lambda x: -x[1]), start=2):
        ws2.cell(row=i, column=1, value=CATEGORY_TR.get(k, k))
        ws2.cell(row=i, column=2, value=round(v, 2))
    ws2.column_dimensions['A'].width = 22; ws2.column_dimensions['B'].width = 18

    ws3 = wb.create_sheet("Aylik Trend")
    hdr(ws3, 1, 1, "Ay"); hdr(ws3, 1, 2, "Gelir (TL)"); hdr(ws3, 1, 3, "Gider (TL)")
    for i, row in enumerate(result.get('monthly_trend', []), start=2):
        ws3.cell(row=i, column=1, value=row.get('month', ''))
        ws3.cell(row=i, column=2, value=row.get('income', 0))
        ws3.cell(row=i, column=3, value=row.get('expense', 0))
    for col in ['A', 'B', 'C']:
        ws3.column_dimensions[col].width = 18

    ws4 = wb.create_sheet("Karar Aciklamasi")
    hdr(ws4, 1, 1, "Faktor"); hdr(ws4, 1, 2, "Deger"); hdr(ws4, 1, 3, "Birim"); hdr(ws4, 1, 4, "Etki")
    for i, f in enumerate(result.get('shap_features', []), start=2):
        ws4.cell(row=i, column=1, value=f.get('display', ''))
        ws4.cell(row=i, column=2, value=f.get('value', 0))
        ws4.cell(row=i, column=3, value=f.get('unit', ''))
        ws4.cell(row=i, column=4, value="Riski Artirdi" if f.get('direction') == 'risk' else (
                                         "Riski Azaltti" if f.get('direction') == 'safe' else "Notr"))
    for col in ['A', 'B', 'C', 'D']:
        ws4.column_dimensions[col].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.post("/pdf")
async def generate_pdf(req: ReportRequest):
    try:
        data = _generate_pdf(req.result, req.mode, req.persona_name)
    except RuntimeError as e:
        raise HTTPException(503, str(e))
    except Exception as e:
        raise HTTPException(500, f"PDF olusturulamadi: {e}")

    now = datetime.now().strftime("%Y%m%d_%H%M")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=FinWise_Rapor_{now}.pdf"},
    )


@router.post("/excel")
async def generate_excel(req: ReportRequest):
    try:
        data = _generate_excel(req.result, req.mode)
    except RuntimeError as e:
        raise HTTPException(503, str(e))
    except Exception as e:
        raise HTTPException(500, f"Excel olusturulamadi: {e}")

    now = datetime.now().strftime("%Y%m%d_%H%M")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=FinWise_Sonuclar_{now}.xlsx"},
    )
